import argparse
import openpyxl
import os
import shutil
import tempfile
import zipfile

from AccessControl.SecurityManagement import newSecurityManager
from DateTime import DateTime
from plone.dexterity.interfaces import IDexterityFTI
from Products.Archetypes import Field
from Products.CMFCore.utils import getToolByName
import zope.interface
from zope.component import getUtility
from zope.dottedname.resolve import resolve
from zope.schema import getFieldsInOrder

try:
    from bika.lims.interfaces import IProxyField
    PROXY_FIELD_INSTALLED = True
except:
    PROXY_FIELD_INSTALLED = False


# def excepthook(typ, value, tb):
#     import pudb as pdb
#     import traceback
#     traceback.print_exception(typ, value, tb)
#     pdb.pm()
#     pdb.set_trace()
# import sys; sys.excepthook = excepthook

export_types = [
    'ReferenceDefinition',
    'ClientDepartment',
    'ClientType',
    'Client',
    'Contact',
    'ARPriority',
    'AnalysisProfile',
    'ARTemplate',
    'AnalysisCategory',
    'AnalysisService',
    'AnalysisSpec',
    'AttachmentType',
    'BatchLabel',
    'Calculation',
    'Container',
    'ContainerType',
    'Department',
    'Instrument',
    'InstrumentCalibration',
    'InstrumentCertification',
    'InstrumentMaintenanceTask',
    'InstrumentScheduledTask',
    'InstrumentType',
    'InstrumentValidation',
    'LabContact',
    'LabProduct',
    'Manufacturer',
    'Method',
    'Preservation',
    'SampleCondition',
    'SampleMatrix',
    'StorageLocation',
    'SamplePoint',
    'SampleType',
    'Strain',
    'SamplingDeviation',
    'SRTemplate',
    'SubGroup',
    #'Supplier',
    #'SupplierContact',
    'UnitConversion',
    'WorksheetTemplate',
    #'ARReport',
    #'Analysis',
    #'AnalysisRequest',
    #'Attachment',
    #'Batch',
    #'BatchFolder',
    'Calculations',
    'ClientFolder',
    'DuplicateAnalysis',
    #'Invoice',
    #'InvoiceBatch',
    'Pricelist',
    'ReferenceAnalysis',
    'ReferenceSample',
    'RejectAnalysis',
    #'Sample',
    #'SamplePartition',
    #'SupplyOrder',
    #'SupplyOrderItem',
    #'Worksheet'
]

# fieldnames that are never exported
ignore_fields = [
    # dublin
    'constrainTypesMode',
    'locallyAllowedTypes',
    'immediatelyAddableTypes',
    'subject',
    'relatedItems',
    'location',
    'language',
    'effectiveDate',
    'modification_date',
    'expirationDate',
    'creators',
    'contributors',
    'rights',
    'allowDiscussion',
    'excludeFromNav',
    'nextPreviousEnabled',
]

app = app  # flake8: noqa

class Main:
    def __init__(self, args):
        self.args = args
        # pose as user
        self.user = app.acl_users.getUserById(args.username)
        newSecurityManager(None, self.user)
        # get portal object
        self.portal = app.unrestrictedTraverse(args.sitepath)

        self.proxy_cache = {}

    def __call__(self):
        """Export entire bika site
        """
        self.tempdir = tempfile.mkdtemp()
        # Export into tempdir
        self.wb = openpyxl.Workbook()
        self.export_laboratory()
        self.export_bika_setup()
        for portal_type in export_types:
            self.export_portal_type(portal_type)
        self.wb.save(os.path.join(self.tempdir, 'setupdata.xlsx'))
        # Create zip file
        zf = zipfile.ZipFile(self.args.outputfile, 'w', zipfile.ZIP_DEFLATED)
        for fname in os.listdir(self.tempdir):
            zf.write(os.path.join(self.tempdir, fname), fname)
        zf.close()
        # Remove tempdir
        shutil.rmtree(self.tempdir)

    def get_catalog(self, portal_type):
        # grab the first catalog we are indexed in
        at = getToolByName(self.portal, 'archetype_tool')
        return at.getCatalogsByType(portal_type)[0]

    def get_fields(self, instance):
        fields = []
        type_info = instance.getTypeInfo()
        if type_info.content_meta_type == 'Dexterity Item':
            field_names = []
            if 'plone.app.dexterity.behaviors.metadata.IBasic' in \
                    type_info.behaviors:
                fields.append(
                    {'id': 'id',
                     'type': '',
                     'value': instance.id})
                fields.append(
                    {'id': 'title',
                     'type': '',
                     'value': instance.title})
                fields.append(
                    {'id': 'description',
                     'type': '',
                     'value': instance.description})
            iface = resolve(type_info.schema)
            for field in getFieldsInOrder(iface):
                fields.append(
                    {'id': field[0], 'type': field[1],
                     'value': instance.get(field[0])})
        else:
            schema = instance.schema
            for field in schema.fields():
                if field.getName() in ignore_fields:
                    continue
                if Field.IComputedField.providedBy(field):
                    continue
                fields.append(field)
        return fields

    def write_dict_field_values(self, instance, field):
        value = field.get(instance)
        if type(value) == dict:
            value = [value]
        #Ensure keys are found for the longest dict
        keys = []
        for val in value:
            for key in val.keys():
                if key not in keys:
                    keys.append(key)
        keys = sorted(keys)

        #Remove special fields that are filled 'manually'
        for k in ('id', 'field'):
            if k in keys:
                keys.remove(k)
        # Create or obtain sheet for this field type's values
        sheetname = '%s_values' % field.type
        sheetname = sheetname[:31]
        if sheetname in self.wb:
            ws = self.wb[sheetname]
        else:
            ws = self.wb.create_sheet(title=sheetname)
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.cell(column=1, row=1).value = "id"
            ws.cell(column=2, row=1).value = "field"
            for col, key in enumerate(keys):
                cell = ws.cell(column=col + 3, row=1)
                cell.value = key
            if field.getName() == 'Analyses':
                col = 3 + len(keys)
            	ws.cell(column=col, row=1).value = "service_id"
            elif field.getName() == 'ReferenceResults':
                col = 3 + len(keys)
            	ws.cell(column=col, row=1).value = "uid"
            elif field.getName() == 'Licenses':
                col = 3 + len(keys)
            	ws.cell(column=col, row=1).value = "client_type"
        nr_rows = len(list(ws.rows)) + 1
        for row, v in enumerate(value):
            if not any(v.values()):
                break
            # source id/field
            ws.cell(column=1, row=nr_rows + row).value = instance.id
            ws.cell(column=2, row=nr_rows + row).value = field.getName()
            for col, key in enumerate(keys):
                c_value = v.get(key, '')
                if hasattr(c_value, 'portal_type'):
                    print 'WARNING: attempt to write instance %s:%s to field %s of instance %s:%s' % (c_value.portal_type, c_value.id, field.getName(), instance.portal_type, instance.id)
                    #import pdb; pdb.set_trace()
                    #c_value = c_value.getId()
                    continue

                try:
                    ws.cell(column=col + 3, row=nr_rows + row).value = c_value
                except Exception, e:
                    print 'Error on %s: %s' % (
                            field.getName(), str(e))
                    import pdb; pdb.set_trace()
                    raise

            if field.getName() == 'Licenses' and v.get('LicenseType', False):
                col = 3 + len(keys)
                types = instance.bika_setup_catalog(UID=v['LicenseType'])
                title = ''
                if len(types):
                    title = types[0].getId
            	ws.cell(column=col, row=nr_rows+row).value = title
            if field.getName() == 'Analyses' and v.get('service_uid', False):
                col = 3 + len(keys)
                services = instance.bika_setup_catalog(UID=v['service_uid'])
                title = ''
                if len(services):
                    title = services[0].getId
            	ws.cell(column=col, row=nr_rows+row).value = title
            if field.getName() == 'ReferenceResults' and \
               v.get('uid', False):
                col = 3 + len(keys)
                services = instance.bika_setup_catalog(UID=v['uid'])
                title = ''
                if len(services):
                    title = services[0].getId
            	ws.cell(column=col, row=nr_rows+row).value = title
        return sheetname

    def write_reference_values(self, instance, field):
        values = field.get(instance)
        # Create or obtain sheet for this relationship
        sheetname = field.relationship[:31]
        if sheetname in self.wb:
            ws = self.wb[sheetname]
        else:
            ws = self.wb.create_sheet(title=sheetname)
            ws.cell(column=1, row=1).value = "Source"
            ws.cell(column=2, row=1).value = "Target"
        nr_rows = len(list(ws.rows)) + 1
        for row, value in enumerate(values):
            ws.cell(column=1, row=nr_rows + row).value = instance.id
            ws.cell(column=2, row=nr_rows + row).value = value.id
        return sheetname

    def get_extension(self, mimetype):
        """Return first extension for mimetype, if any is found.
        If no extension found, return ''
        """
        mr = getToolByName(self.portal, "mimetypes_registry")
        extension = ''
        for ext, mt in mr.extensions.items():
            if mimetype == mt:
                extension = ext
        return extension

    def mutate(self, instance, field):
        #if instance.portal_type == "ARTemplate" and \
        #   field.getName() == 'Analyses':
        if type(field) == dict:
            #Dexterity
            return getattr(instance, field['id'])

        value = field.get(instance)
        # Booleans are special; we'll str and return them.
        if value is True or value is False:
            return str(value)
        # Zero is special: it's false-ish, but the value is important.
        if value is 0:
            return 0
        # Other falsish values make empty cells.
        if not value:
            return ''
        # Ignore Analyses fields
        if instance.portal_type != "ARTemplate" and \
           field.getName() == 'Analyses':
            print 'Ignore Analysis field for %s: %s' % (
                instance.portal_type, instance.getId())
            return None
        #### Ignore ProxyFields
        if PROXY_FIELD_INSTALLED and IProxyField.providedBy(field):
            return None
        # Date fields get stringed to rfc8222
        if Field.IDateTimeField.providedBy(field):
            return value.rfc822() if value else None
        # TextField implements IFileField, so we must handle it
        # before IFileField. It's just returned verbatim.
        elif Field.ITextField.providedBy(field):
            return value
        # Files get saved into tempdir, and the cell content is the filename
        elif Field.IFileField.providedBy(field):
            if not value.size:
                return ''
            extension = self.get_extension(value.content_type)
            filename = value.filename if value.filename \
                else instance.id + '-' + field.getName() + "." + extension
            of = open(os.path.join(self.tempdir, filename), 'wb')
            try:
                of.write(value.data.data)
            except:
                try:
                    of.write(value.data)
                except:
                    import pdb; pdb.set_trace()
            of.close()
            return filename
        elif Field.IReferenceField.providedBy(field):
            if field.multiValued:
                return self.write_reference_values(instance, field)
            else:
                return value.id
        elif Field.ILinesField.providedBy(field):
            return "\n".join(value)
        # depend on value of field, to decide mutation.
        else:
            value = field.get(instance)
            # Dictionaries or lists of dictionaries
            if type(value) == dict and value == {}:
                return ''
            elif type(value) in (list, tuple) and len(value) == 0:
                return ''
            elif type(value) == dict \
                    or (type(value) in (list, tuple)
                        and len(value) and type(value[0]) == dict):
                return self.write_dict_field_values(instance, field)
            else:
                return value

    def export_laboratory(self):
        instance = self.portal.bika_setup.laboratory
        ws = self.wb.create_sheet(title='Laboratory')
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        fields = self.get_fields(instance)
        for row, field in enumerate(fields):
            ws.cell(column=1, row=row + 1).value = field.getName()
            value = self.mutate(instance, field)
            ws.cell(column=2, row=row + 1).value = value

    def export_bika_setup(self):
        instance = self.portal.bika_setup
        ws = self.wb.create_sheet(title='BikaSetup')
        fields = self.get_fields(instance)
        for row, field in enumerate(fields):
            ws.cell(column=1, row=row + 1).value = field.getName()
            value = self.mutate(instance, field)
            ws.cell(column=2, row=row + 1).value = value

    def export_portal_type(self, portal_type):
        def get_headers(fields):
            headers = []
            for field in fields:
                if type(field) == dict:
                    headers.append(field['id'])
                else:
                    headers.append(field.getName())
            return headers

        catalog = self.get_catalog(portal_type)
        brains = catalog(portal_type=portal_type)
        if not brains:
            print "No objects of type %s found in %s" % (portal_type, catalog)
            return
        ws = self.wb.create_sheet(title=portal_type)
        # Write headers
        instance = brains[0].getObject()
        fields = self.get_fields(instance)
        headers = ['path', 'uid']
        headers += get_headers(fields)
        for col, header in enumerate(headers):
            ws.cell(column=col + 1, row=1).value = header
        # Write values
        portal_path = '/'.join(self.portal.getPhysicalPath())
        
        for row, brain in enumerate(brains):
            instance = brain.getObject()
            # path
            path = '/'.join(instance.getPhysicalPath()[:-1])
            ws.cell(column=1, row=row + 2).value = \
                path.replace(portal_path, '')
            # uid
            ws.cell(column=2, row=row + 2).value = instance.UID()
            # then schema field values
            for col, field in enumerate(fields):
                value = self.mutate(instance, field)
                try:
                    #print 'Set Cell (%s, %s) to %s' % ( col+3, row+2, value)
                    ws.cell(column=col + 3, row=row + 2).value = value
                except Exception, e:
                    print 'Error on %s: %s' % (
                            field.getName(), str(e))
                    import pdb; pdb.set_trace()
                    raise


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Export bika_setup into an Open XML (XLSX) workbook',
        epilog='This script is meant to be run with zopepy or bin/instance.'
               'See http://docs.plone.org/develop/plone/misc/commandline.html'
               'for details.'
    )
    parser.add_argument(
        '-s',
        dest='sitepath',
        default='Plone',
        help='full path to site root (default: Plone)')
    parser.add_argument(
        '-u',
        dest='username',
        default='admin',
        help='zope admin username (default: admin)')
    parser.add_argument(
        '-o',
        dest='outputfile',
        default='',
        help='output zip file name (default: SITEPATH.zip)')
    args, unknown = parser.parse_known_args()
    if args.outputfile == '':
        args.outputfile = args.sitepath + ".zip"

    main = Main(args)
    main()
